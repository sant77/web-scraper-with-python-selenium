from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import Progress
from rich.style import Style
from rich.box import ROUNDED
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from datetime import datetime
from src.utils.logs_config import set_logs_configuration
from selenium.webdriver.common.action_chains import ActionChains
import logging
import platform
import random
import re
import time
import requests
from dotenv import load_dotenv
import os

class ScraperSelenium():

    def __init__(self, webpage_url:str, two_captcha_api_key:str = None) -> None:
        self.webpage_url = webpage_url
        self.pattern_nit = r'[^0-9]'
        self.date_pattern = r"\b(19\d\d|20\d\d)[-/](0[1-9]|1[0-2])[-/](0[1-9]|[12]\d|3[01])\b"
        self.attempts = 3  # Reducido porque ahora usaremos 2Captcha
        self.two_captcha_api_key = two_captcha_api_key
        self.cloudflare_sitekey = "0x4AAAAAAAg1WuNb-OnOa76z"  # Sitekey de Cloudflare Turnstile

    def solve_captcha_with_2captcha(self):
        """Resuelve el captcha usando el servicio 2Captcha"""
        if not self.two_captcha_api_key:
            raise ValueError("API key de 2Captcha no proporcionada")
        
        # URL de la página para obtener el parámetro data-action
        page_url = self.webpage_url
        
        # Enviamos la solicitud a 2Captcha
        s = requests.Session()
        captcha_params = {
            'key': self.two_captcha_api_key,
            'method': 'turnstile',
            'sitekey': self.cloudflare_sitekey,
            'pageurl': page_url,
            'json': 1
        }
        
        try:
            # Enviar la solicitud para resolver el captcha
            response = s.post('http://2captcha.com/in.php', data=captcha_params)
            request_id = response.json().get('request')
            
            if not request_id:
                logging.error(f"Error al enviar captcha a 2Captcha: {response.text}")
                return None
                
            # Esperar la solución
            for _ in range(30):  # Máximo 30 intentos (30 segundos)
                time.sleep(1)  # Esperar 1 segundo entre intentos
                result = s.get(f'http://2captcha.com/res.php?key={self.two_captcha_api_key}&action=get&id={request_id}&json=1').json()
                
                if result.get('status') == 1:
                    return result.get('request')  # Token de solución
                    
            logging.error("Tiempo de espera agotado para 2Captcha")
            return None
            
        except Exception as e:
            logging.error(f"Error al resolver captcha con 2Captcha: {e}")
            return None

    def find_dane_elements(self, cufe:str):
        set_logs_configuration()
       
        if platform.system() == "Linux":
            chrome_options = Options()
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--remote-debugging-port=9222")
            chrome_options.binary_location = "/opt/chrome/chrome-linux64/chrome"
            chrome_service = ChromeService(executable_path="/opt/chromedriver/chromedriver-linux64/chromedriver")
            driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
        else:
            driver = webdriver.Chrome() 

        try:
            driver.get(self.webpage_url)
            time.sleep(random.randint(1, 3))
            
            # Ingresar el CUFE
            input_text_fname = driver.find_element(By.ID, 'DocumentKey')
            input_text_fname.send_keys(cufe)
            
            # Resolver el captcha con 2Captcha
            captcha_token = self.solve_captcha_with_2captcha()
            if not captcha_token:
                logging.error("No se pudo resolver el captcha con 2Captcha")
                driver.quit()
                return None, None, None, None, None, None
                
            # Inyectar el token de solución en la página
            try:
                wait = WebDriverWait(driver, 10)
                captcha_element = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'input[name="cf-turnstile-response"]'))
                )
                driver.execute_script(f'arguments[0].value = "{captcha_token}";', captcha_element)
            except Exception as e:
                logging.error(f"No se pudo encontrar/inyectar el elemento del captcha: {e}")
                raise
            
            # Hacer clic en el botón de búsqueda
            button = driver.find_element(By.CLASS_NAME, 'btn.btn-primary.search-document.margin-top-40')
            button.click()
            
            # Esperar a que la página cargue (máximo 10 segundos)
            for _ in range(10):
                if driver.current_url != self.webpage_url:
                    break
                time.sleep(1)
            else:
                logging.warn("No se pudo pasar la página de captcha")
                driver.quit()
                return None, None, None, None, None, None

            # Obtener los datos de la página de resultados
            text = driver.find_elements(By.CLASS_NAME, 'col-md-4')
            table = driver.find_elements(By.CLASS_NAME, "table-responsive")
            download = driver.find_element(By.CLASS_NAME, "downloadPDFUrl")
            link = download.get_property('href')
        
            emisor_list = text[1].text.split(" ") 
            receptor_list = text[2].text.split(" ")

            emisor_nit = re.sub(self.pattern_nit, "", emisor_list[3])
            receptor_nit = re.sub(self.pattern_nit, "", receptor_list[3])

            emisor_name = self._join_name(emisor_list[4:])
            receptor_name = self._join_name(receptor_list[4:])

            events = self._look_for_events(table[1].text.split())
            
            driver.quit()
            return link, emisor_nit, emisor_name, receptor_nit, receptor_name, events
        
        except Exception as e:
            logging.error(f"Error inesperado en el scraping: {e}")
            driver.quit()
            raise Exception(e)
            
    def _look_for_events(self, table:list):
        events = []
        state0 = 1
        state1 = 0
        state2 = 0
        patron_numeros = r'^\d+$'
        number_event = ""
        type_of_event = ""

        for i in table[9:]:
            num = re.match(patron_numeros, i)

            if num != None and state0 == 1:
                number_event = i
                state1 = 1
                state0 = 0
                state2 = 0

            if state1 == 1 and i != number_event:
                if re.search(self.date_pattern, i):
                    state2 = 1
                    state1 = 0
                    state0 = 0
                else:
                    type_of_event += i + " "
            
            if state2 == 1 and i == "detalle":
                events.append({
                    "eventNumber": number_event,
                    "eventName": type_of_event
                })
                number_event = ""
                type_of_event = ""
                state0 = 1
                state1 = 0
                state2 = 0

        return events
        
    def _join_name(self, name:list):
        return " ".join(name)
    
    def export_to_excel(self, data: List[Dict], filename: str = "dian_results.xlsx") -> str:
        """
        Exporta los datos a un archivo Excel con formato profesional
        
        Args:
            data: Lista de diccionarios con los datos a exportar
            filename: Nombre del archivo de salida
            
        Returns:
            Ruta completa del archivo generado
        """
        # Crear el libro de trabajo y la hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados DIAN"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        # Encabezados
        headers = [
            "N°", "Emisor NIT", "Emisor Nombre", 
            "Receptor NIT", "Receptor Nombre", 
            "Enlace PDF", "N° Evento", "Tipo de Evento"
        ]
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Escribir datos
        row_num = 2
        for idx, item in enumerate(data, start=1):
            # Datos principales
            main_data = [
                idx,
                item.get("emisor_nit", "N/A"),
                item.get("emisor_name", "N/A"),
                item.get("receptor_nit", "N/A"),
                item.get("receptor_name", "N/A"),
                item.get("link", "N/A")
            ]
            
            # Escribir fila principal
            for col_num, value in enumerate(main_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
            
            # Escribir eventos si existen
            if item.get("events"):
                for event in item["events"]:
                    ws.cell(row=row_num, column=7, value=event.get("eventNumber", "N/A")).border = thin_border
                    ws.cell(row=row_num, column=8, value=event.get("eventName", "N/A")).border = thin_border
                    row_num += 1
            else:
                ws.cell(row=row_num, column=7, value="Sin eventos").border = thin_border
                ws.cell(row=row_num, column=8, value="N/A").border = thin_border
                row_num += 1
        
        # Ajustar el ancho de las columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Congelar paneles (filas de encabezado)
        ws.freeze_panes = "A2"
        
        # Añadir filtros
        ws.auto_filter.ref = ws.dimensions
        
        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        final_filename = f"dian_results_{timestamp}.xlsx"
        
        # Guardar el archivo
        wb.save(final_filename)
        
        return os.path.abspath(final_filename)
    
    def display_results(self, data: List[Dict]) -> None:
        """Muestra los resultados en una tabla formateada con Rich"""
        console = Console()
        
        # Estilos personalizados
        success_style = Style(color="green", bold=True)
        warning_style = Style(color="yellow", bold=True)
        error_style = Style(color="red", bold=True)
        info_style = Style(color="blue", bold=True)
        
        console.print(Panel.fit("Resultados del Scraping DIAN", 
                              style=info_style, 
                              border_style=success_style))
        
        for idx, result in enumerate(data, start=1):
            # Panel para cada resultado
            console.print(Panel.fit(f"Resultado #{idx} - CUFE procesado", 
                           style=info_style))
            
            # Tabla de información principal
            main_table = Table(show_header=True, header_style="bold magenta", 
                             box=ROUNDED, expand=True)
            main_table.add_column("Campo", style="cyan", width=20)
            main_table.add_column("Valor", style="white")
            
            main_table.add_row("Emisor NIT", result.get("emisor_nit", "N/A"))
            main_table.add_row("Emisor Nombre", result.get("emisor_name", "N/A"))
            main_table.add_row("Receptor NIT", result.get("receptor_nit", "N/A"))
            main_table.add_row("Receptor Nombre", result.get("receptor_name", "N/A"))
            main_table.add_row("Enlace PDF", result.get("link", "N/A"))
            
            console.print(main_table)
            
            # Tabla de eventos si existen
            if result.get("events"):
                events_table = Table(show_header=True, header_style="bold yellow", 
                                   box=ROUNDED, title="Eventos")
                events_table.add_column("Número", style="cyan")
                events_table.add_column("Tipo de Evento", style="white")
                
                for event in result["events"]:
                    events_table.add_row(event.get("eventNumber", "N/A"), 
                                       event.get("eventName", "N/A"))
                
                console.print(events_table)
            else:
                console.print("[yellow]No se encontraron eventos[/yellow]")
            
            console.print()  # Espacio entre resultados

if __name__ == "__main__":
    # Ejemplo de uso con 2Captcha
    load_dotenv()
    API_KEY_2CAPTCHA = os.environ.get("API_KEY_2CAPTCHA")

    scraper = ScraperSelenium(
        "https://catalogo-vpfe.dian.gov.co/User/SearchDocument",
        two_captcha_api_key=API_KEY_2CAPTCHA
    )

    cufes = [
        "2320c5d9dcb19b24b9e8baa0e033579a1b1329bb1c846850838c8b6daedb90d36f528a727a79041869b68b7c2f2352c0",
        "1f28b0cafdafdfc493c2d2abff1168fe99f56395f8c77f7ae492c31972c404ddc54339e51cad28e7e77277a44ca3664e",
        "86da9212194f131522f9f450e623ae11086fdc5f782f58706229a71663af58cd11d231de7ad2b8cb13f2d12a263a5c9c",
        "2320c5d9dcb19b24b9e8baa0e033579a1b1329bb1c846850838c8b6daedb90d36f528a727a79041869b68b7c2f2352c0",
        "fa7f59b23c5d77ce97808a2b5628b81da05b86eb1aa713d3bddbecb559070c6a45cbcba215161f755826805c353bec8c",
        "728c9ac2c2cfeb509825f76adc4fce9964142aac8dcfcbda66785013f8d0d93e8d6603857bd49885c296dce70856b227",
        "71546c618e7464e422088512e32c50f3ff96e80777ca273f98fd1ff1b5c726da1f0c4f9246b2807b69fe4ec60d92336f"
    ]

    data = []

    console = Console()
    with Progress() as progress:
        task = progress.add_task("[cyan]Procesando CUFE...", total=len(cufes))
        
        for cufe in cufes:
            progress.update(task, advance=1, description=f"[cyan]Procesando {cufe[:10]}...")
            
            try:
                link, emisor_nit, emisor_name, receptor_nit, receptor_name, events = scraper.find_dane_elements(cufe)
                
                data.append({
                    "link": link, 
                    "emisor_nit": emisor_nit, 
                    "emisor_name": emisor_name, 
                    "receptor_nit": receptor_nit, 
                    "receptor_name": receptor_name, 
                    "events": events
                })
                
            except Exception as e:
                console.print(f"[red]Error procesando CUFE {cufe[:10]}...: {str(e)}[/red]")
                data.append({
                    "link": None, 
                    "emisor_nit": None, 
                    "emisor_name": None, 
                    "receptor_nit": None, 
                    "receptor_name": None, 
                    "events": None
                })
    
    # Mostrar resultados con formato mejorado
    scraper.display_results(data)

    # Exportar a Excel
    excel_path = scraper.export_to_excel(data)
    console.print(f"\n[bold green]✓ Datos exportados a Excel: [underline]{excel_path}[/underline][/bold green]")